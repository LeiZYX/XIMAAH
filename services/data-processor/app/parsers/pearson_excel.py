from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook

from app.models import TimetableRow


def _excel_serial_to_iso(value) -> str | None:
    if isinstance(value, (int, float)) and value:
        base = date(1899, 12, 30)
        parsed = base + timedelta(days=float(value))
        return parsed.isoformat()

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, str) and value.strip():
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).date().isoformat()
        except ValueError:
            pass
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date().isoformat()
            except ValueError:
                continue

    return None


def _parse_duration(value) -> int | None:
    text = str(value or "").strip()
    match_hm = __import__("re").search(r"(\d+)\s*h\s*(\d+)\s*m", text, __import__("re").I)
    if match_hm:
        return int(match_hm.group(1)) * 60 + int(match_hm.group(2))
    return None


def _normalize_paper_code(raw_code: str) -> tuple[str, str]:
    cleaned = " ".join(raw_code.strip().split())
    parts = cleaned.split(" ")
    if len(parts) >= 2:
        syllabus = parts[0]
        component = "".join(parts[1:])
        return syllabus, f"{syllabus}/{component}"
    return cleaned, cleaned


def _session_start_time(slot: str) -> str:
    normalized = slot.strip().lower()
    if "afternoon" in normalized or "pm" in normalized:
        return "13:30"
    return "09:00"


def _header_index(headers: list[str], names: list[str]) -> int:
    lower = [header.strip().lower() for header in headers]
    for name in names:
        idx = lower.index(name.lower()) if name.lower() in lower else -1
        if idx >= 0:
            return idx
    return -1


def parse_pearson_excel(content: bytes, filename: str = "timetable.xlsx") -> list[TimetableRow]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet_name = next(
        (
            name
            for name in workbook.sheetnames
            if name.lower() == "all papers" or "all" in name.lower()
        ),
        workbook.sheetnames[0],
    )
    sheet = workbook[sheet_name]
    matrix = [list(row) for row in sheet.iter_rows(values_only=True)]

    header_row_index = next(
        (
            index
            for index, row in enumerate(matrix)
            if any(str(cell or "").strip().lower() == "examination code" for cell in row)
        ),
        -1,
    )
    if header_row_index < 0:
        raise ValueError('Could not find "All papers" sheet with examination data')

    headers = [str(cell or "").strip() for cell in matrix[header_row_index]]
    date_idx = _header_index(headers, ["date"])
    qual_idx = _header_index(headers, ["qual", "qualification"])
    code_idx = _header_index(headers, ["examination code", "code"])
    subject_idx = _header_index(headers, ["subject"])
    title_idx = _header_index(headers, ["title"])
    time_idx = _header_index(headers, ["time"])
    duration_idx = _header_index(headers, ["duration"])

    if date_idx < 0 or code_idx < 0 or subject_idx < 0:
        raise ValueError("Pearson Edexcel timetable format not recognised")

    rows: list[TimetableRow] = []
    for row in matrix[header_row_index + 1 :]:
        examination_code = str(row[code_idx] or "").strip()
        subject = str(row[subject_idx] or "").strip()
        if not examination_code or not subject:
            continue

        parsed_date = _excel_serial_to_iso(row[date_idx])
        if not parsed_date:
            continue

        syllabus_code, paper_code = _normalize_paper_code(examination_code)
        qualification_level = (
            str(row[qual_idx] or "").strip() or "GCSE" if qual_idx >= 0 else "GCSE"
        )

        rows.append(
            TimetableRow(
                date=parsed_date,
                qualification_level=qualification_level,
                syllabus_code=syllabus_code,
                paper_code=paper_code,
                subject=subject,
                title=str(row[title_idx] or paper_code).strip() if title_idx >= 0 else paper_code,
                start_time=_session_start_time(str(row[time_idx] or "")) if time_idx >= 0 else "09:00",
                duration_minutes=_parse_duration(row[duration_idx]) if duration_idx >= 0 else None,
            )
        )

    if not rows:
        raise ValueError(f"No exam rows found in {filename}")

    return rows
